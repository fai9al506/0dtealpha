"""
Export ES Absorption signals to Excel with swing pair details for ATAS cross-check.
Each row = one ES Absorption trade.
Columns: trade info + Swing 1 (ref) + Swing 2 (recent) + Trigger Bar details.
"""
import psycopg2, os, json
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

conn = psycopg2.connect(os.environ['DATABASE_URL'])
cur = conn.cursor()

# Get ALL ES Absorption trades with abs_details
cur.execute("""
SELECT s.id, s.ts, s.setup_name, s.direction, s.grade, s.score, s.spot,
       s.outcome_result, s.outcome_pnl, s.outcome_max_profit, s.outcome_max_loss,
       s.paradigm, s.abs_es_price, s.abs_details::text, s.comments,
       s.abs_vol_ratio
FROM setup_log s
WHERE s.setup_name = 'ES Absorption'
ORDER BY s.id
""")
trades = cur.fetchall()
print(f"Found {len(trades)} ES Absorption trades")

wb = Workbook()
ws = wb.active
ws.title = "ES Absorption Signals"

# Header styles
header_font = Font(bold=True, color="FFFFFF", size=10)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
swing1_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # green
swing2_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # blue
trigger_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # orange
win_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Define columns
headers = [
    # Trade info
    "ID", "Date", "Time (ET)", "Direction", "Pattern", "Grade", "Score",
    "SPX Spot", "ES Price", "Paradigm",
    # Outcome
    "Result", "PnL", "Max Profit", "Max Loss",
    # Ref Swing (older / Swing 1)
    "Sw1 Type", "Sw1 Time (ET)", "Sw1 Bar#", "Sw1 Price", "Sw1 CVD", "Sw1 Volume",
    # Recent Swing (Swing 2)
    "Sw2 Type", "Sw2 Time (ET)", "Sw2 Bar#", "Sw2 Price", "Sw2 CVD", "Sw2 Volume",
    # Divergence metrics
    "Price Δ", "CVD Δ", "CVD Z-score", "Price ATR", "Vol Ratio", "ATR",
    # Trigger bar
    "Trig Bar#", "Trig Time (ET)", "Trig Open", "Trig High", "Trig Low", "Trig Close",
    "Trig Volume", "Trig Buy Vol", "Trig Sell Vol", "Trig Delta", "Trig CVD",
    # Extra
    "Tier", "Resolution", "Comments"
]

# Write headers
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# Color group headers
for col in range(15, 21):  # Sw1 columns
    ws.cell(row=1, column=col).fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
for col in range(21, 27):  # Sw2 columns
    ws.cell(row=1, column=col).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
for col in range(31, 42):  # Trigger columns
    ws.cell(row=1, column=col).fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")

row_num = 2
for t in trades:
    tid, ts, setup_name, direction, grade, score, spot, \
        result, pnl, max_profit, max_loss, paradigm, es_price, \
        abs_details_str, comments, vol_ratio_db = t

    abs_d = json.loads(abs_details_str) if abs_details_str else {}
    best = abs_d.get('best_swing', {})
    swing2 = best.get('swing', {})   # recent swing
    swing1 = best.get('ref_swing', {})  # older/reference swing
    pattern = abs_d.get('pattern', '')
    cvd_z = best.get('cvd_z', '')
    price_atr = best.get('price_atr', '')
    vol_ratio = abs_d.get('vol_ratio', '')
    atr = abs_d.get('atr', '')
    trigger_bar_idx = abs_d.get('bar_idx', None)
    pattern_tier = abs_d.get('pattern_tier', '')
    resolution = abs_d.get('resolution_reason', '')

    # Convert timestamps to ET
    ts_et = ts - timedelta(hours=5) if ts else None
    date_str = ts_et.strftime('%Y-%m-%d') if ts_et else ''
    time_str = ts_et.strftime('%H:%M:%S') if ts_et else ''

    def parse_ts_to_et(ts_str):
        """Parse various timestamp formats to ET string"""
        if not ts_str:
            return ''
        from datetime import datetime
        try:
            # Try ISO format with timezone
            if '+' in str(ts_str) or 'Z' in str(ts_str):
                # Already has TZ info
                from dateutil import parser as dtparser
                dt = dtparser.parse(str(ts_str))
                # Convert to ET (UTC-5)
                import pytz
                et = pytz.timezone('US/Eastern')
                dt_et = dt.astimezone(et)
                return dt_et.strftime('%H:%M:%S')
            else:
                return str(ts_str)[:8]
        except:
            return str(ts_str)[:19]

    sw1_time = parse_ts_to_et(swing1.get('ts', ''))
    sw2_time = parse_ts_to_et(swing2.get('ts', ''))

    # Calculate price delta and CVD delta between swings
    sw1_price = swing1.get('price', 0) or 0
    sw2_price = swing2.get('price', 0) or 0
    sw1_cvd = swing1.get('cvd', 0) or 0
    sw2_cvd = swing2.get('cvd', 0) or 0
    price_delta = round(sw2_price - sw1_price, 2) if sw1_price and sw2_price else ''
    cvd_delta = round(sw2_cvd - sw1_cvd, 1) if sw1_cvd is not None and sw2_cvd is not None else ''

    # Get trigger bar data from es_range_bars
    trig_data = {}
    if trigger_bar_idx is not None and ts_et:
        trade_date = ts_et.strftime('%Y-%m-%d')
        cur.execute("""
            SELECT bar_idx, bar_open, bar_high, bar_low, bar_close,
                   bar_volume, bar_buy_volume, bar_sell_volume, bar_delta,
                   cvd_close, ts_start, ts_end
            FROM es_range_bars
            WHERE trade_date = %s AND bar_idx = %s AND source = 'rithmic'
            LIMIT 1
        """, (trade_date, trigger_bar_idx))
        trig_row = cur.fetchone()
        if not trig_row:
            # Try live source
            cur.execute("""
                SELECT bar_idx, bar_open, bar_high, bar_low, bar_close,
                       bar_volume, bar_buy_volume, bar_sell_volume, bar_delta,
                       cvd_close, ts_start, ts_end
                FROM es_range_bars
                WHERE trade_date = %s AND bar_idx = %s AND source = 'live'
                LIMIT 1
            """, (trade_date, trigger_bar_idx))
            trig_row = cur.fetchone()
        if trig_row:
            trig_data = {
                'bar_idx': trig_row[0],
                'open': float(trig_row[1]) if trig_row[1] else '',
                'high': float(trig_row[2]) if trig_row[2] else '',
                'low': float(trig_row[3]) if trig_row[3] else '',
                'close': float(trig_row[4]) if trig_row[4] else '',
                'volume': int(trig_row[5]) if trig_row[5] else '',
                'buy_vol': int(trig_row[6]) if trig_row[6] else '',
                'sell_vol': int(trig_row[7]) if trig_row[7] else '',
                'delta': int(trig_row[8]) if trig_row[8] else '',
                'cvd': float(trig_row[9]) if trig_row[9] else '',
                'ts_start': trig_row[10],
            }
            trig_data['time_et'] = parse_ts_to_et(trig_data.get('ts_start', ''))

    # Write row
    values = [
        tid, date_str, time_str, direction, pattern, grade, score,
        float(spot) if spot else '', float(es_price) if es_price else '', paradigm or '',
        result or 'OPEN', float(pnl) if pnl else 0, float(max_profit) if max_profit else 0, float(max_loss) if max_loss else 0,
        # Swing 1 (ref/older)
        swing1.get('type', ''), sw1_time, swing1.get('bar_idx', ''),
        sw1_price if sw1_price else '', sw1_cvd if sw1_cvd else '', swing1.get('volume', ''),
        # Swing 2 (recent)
        swing2.get('type', ''), sw2_time, swing2.get('bar_idx', ''),
        sw2_price if sw2_price else '', sw2_cvd if sw2_cvd else '', swing2.get('volume', ''),
        # Divergence
        price_delta, cvd_delta, round(cvd_z, 2) if isinstance(cvd_z, (int, float)) else cvd_z,
        round(price_atr, 2) if isinstance(price_atr, (int, float)) else price_atr,
        round(vol_ratio, 2) if isinstance(vol_ratio, (int, float)) else vol_ratio,
        round(atr, 2) if isinstance(atr, (int, float)) else atr,
        # Trigger bar
        trig_data.get('bar_idx', trigger_bar_idx or ''),
        trig_data.get('time_et', ''),
        trig_data.get('open', ''), trig_data.get('high', ''),
        trig_data.get('low', ''), trig_data.get('close', ''),
        trig_data.get('volume', ''), trig_data.get('buy_vol', ''),
        trig_data.get('sell_vol', ''), trig_data.get('delta', ''),
        trig_data.get('cvd', ''),
        # Extra
        pattern_tier, resolution, (comments or '')[:100]
    ]

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Color result column
    result_cell = ws.cell(row=row_num, column=11)
    if result == 'WIN':
        result_cell.fill = win_fill
    elif result == 'LOSS':
        result_cell.fill = loss_fill

    # Color swing columns
    for col in range(15, 21):
        ws.cell(row=row_num, column=col).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    for col in range(21, 27):
        ws.cell(row=row_num, column=col).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for col in range(31, 42):
        ws.cell(row=row_num, column=col).fill = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")

    row_num += 1

# Auto-width columns
for col in range(1, len(headers) + 1):
    max_len = len(str(ws.cell(row=1, column=col).value or ''))
    for row in range(2, row_num):
        val = str(ws.cell(row=row, column=col).value or '')
        if len(val) > max_len:
            max_len = len(val)
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 3, 25)

# Freeze top row + first 4 columns
ws.freeze_panes = 'E2'

# Add a summary sheet
ws2 = wb.create_sheet("Summary")
ws2.cell(row=1, column=1, value="ES Absorption Signal Summary").font = Font(bold=True, size=14)
ws2.cell(row=3, column=1, value="Column Guide:").font = Font(bold=True)
guide = [
    ("Green columns (Sw1)", "Reference swing (older) — the first swing in the pair"),
    ("Blue columns (Sw2)", "Recent swing — the second swing in the pair"),
    ("Orange columns (Trig)", "Trigger bar — the bar that fired the signal"),
    ("Pattern", "sell_exhaustion / sell_absorption (bullish), buy_exhaustion / buy_absorption (bearish)"),
    ("CVD Z-score", "How many std devs the CVD gap is from normal (higher = stronger divergence)"),
    ("Price ATR", "Price distance between swings as ATR multiple"),
    ("Vol Ratio", "Trigger bar volume / 10-bar rolling avg (>1.4 = elevated)"),
    ("Tier", "2=exhaustion (higher priority), 1=absorption"),
    ("", ""),
    ("HOW TO CROSS-CHECK IN ATAS:", ""),
    ("1.", "Load ES 5-pt range bars (same as our system)"),
    ("2.", "Find each trade by Trig Time (ET) column"),
    ("3.", "Compare Sw1 and Sw2 candle data (price, CVD, volume) with ATAS"),
    ("4.", "Check if the swing highs/lows and CVD values match"),
    ("5.", "Note any discrepancies in the 'Notes' column you can add"),
]
for i, (label, desc) in enumerate(guide, 5):
    ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
    ws2.cell(row=i, column=2, value=desc)
ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 80

# Save
out_path = "ES_Absorption_Signals.xlsx"
wb.save(out_path)
print(f"\nSaved {row_num - 2} trades to {out_path}")

conn.close()
