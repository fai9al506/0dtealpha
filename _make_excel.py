"""Create Excel sheet for V4 signals — Mar 19"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "V4 Signals Mar 19"

# Styles
hdr_fill = PatternFill("solid", fgColor="2F5496")
hdr_font = Font(color="FFFFFF", bold=True, size=10)
thin = Side(style="thin")
border = Border(top=thin, bottom=thin, left=thin, right=thin)
green_fill = PatternFill("solid", fgColor="C6EFCE")
red_fill = PatternFill("solid", fgColor="FFC7CE")
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
user_fill = PatternFill("solid", fgColor="BDD7EE")

headers = [
    "#", "Time ET", "Dir", "Open", "High", "Low", "Close", "Color",
    "Delta", "Body", "Peak Delta", "Trough Delta", "Volume", "Vol/sec",
    "Delta Z", "Grade", "Score", "Outcome", "PnL", "MFE", "MAE",
    "Your Signal?", "Your Comment"
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

# Pre-mark which ones the user identified
user_times = ["09:35", "09:46", "09:50", "09:54", "10:11", "11:05",
              "11:36", "12:07", "12:35", "12:40", "12:42", "12:52",
              "14:13", "14:41", "15:04"]

signals_data = [
    [1,"09:35:01","BULL",6617.50,6619.50,6614.50,6619.50,"GREEN",-324,2.0,2,-445,3250,100.6,0.33,"C",35,"WIN",9.5,17.5,-6.2],
    [2,"09:44:04","BEAR",6634.75,6637.00,6632.00,6632.00,"RED",136,2.75,398,0,2104,90.5,0.32,"C",45,"LOSS",-3.2,4.8,-3.8],
    [3,"09:45:39","BULL",6630.75,6632.25,6627.25,6632.25,"GREEN",-402,1.5,4,-485,2156,123.8,2.21,"A",70,"LOSS",-4.5,3.5,-5.8],
    [4,"09:52:28","BULL",6642.75,6645.00,6640.00,6645.00,"GREEN",-160,2.25,0,-363,4264,102.4,0.52,"B",55,"LOSS",-2.8,5.2,-4.0],
    [5,"10:00:49","BEAR",6649.50,6652.50,6647.50,6647.50,"RED",414,2.0,576,-7,5850,80.2,1.15,"B",60,"LOSS",-1.0,7.0,-5.0],
    [6,"10:04:47","BULL",6644.50,6647.25,6642.25,6647.25,"GREEN",-162,2.75,0,-221,1602,65.1,0.34,"C",35,"LOSS",-5.2,2.8,-6.0],
    [7,"10:06:21","BEAR",6647.25,6647.50,6642.50,6642.50,"RED",307,4.75,310,-18,1711,45.8,0.84,"C",45,"LOSS",-6.0,2.0,-7.5],
    [8,"10:11:35","BEAR",6653.75,6655.50,6650.50,6650.50,"RED",549,3.25,801,-21,7667,78.5,2.14,"A",75,"WIN",9.8,17.8,-1.5],
    [9,"10:16:47","BULL",6638.00,6641.25,6636.25,6641.25,"GREEN",-183,3.25,17,-403,5121,57.3,0.15,"C",45,"LOSS",-7.8,0.2,-7.8],
    [10,"10:31:35","BULL",6630.50,6633.75,6628.75,6633.75,"GREEN",-129,3.25,1,-420,2229,105.2,0.53,"B",60,"WIN",9.8,17.8,-0.8],
    [11,"10:38:58","BEAR",6635.75,6636.25,6631.25,6636.25,"GREEN",159,0.5,162,-248,4251,100.9,0.33,"C",35,"LOSS",-5.0,3.0,-6.5],
    [12,"10:42:59","BULL",6636.25,6640.75,6635.75,6635.75,"RED",-138,0.5,139,-138,1948,38.0,0.57,"LOG",30,"LOSS",-4.5,3.5,-4.5],
    [13,"10:44:23","BEAR",6639.25,6639.25,6634.25,6634.25,"RED",405,5.0,457,-41,3711,41.5,1.18,"B",60,"LOSS",-5.0,3.0,-5.0],
    [14,"10:59:05","BEAR",6639.75,6642.00,6637.00,6637.00,"RED",305,2.75,480,-80,7093,40.0,0.48,"C",45,"LOSS",-6.5,1.5,-7.8],
    [15,"11:23:10","BULL",6635.50,6639.75,6634.75,6634.75,"RED",-545,0.75,252,-545,6883,66.2,0.42,"C",40,"LOSS",-2.0,6.0,-5.0],
    [16,"11:36:19","BULL",6627.25,6629.25,6624.25,6629.25,"GREEN",-552,2.0,24,-955,7500,54.4,0.70,"B",50,"LOSS",-4.0,4.0,-4.2],
    [17,"11:43:42","BEAR",6633.00,6633.25,6628.25,6628.25,"RED",119,4.75,213,0,3045,111.3,0.95,"B",55,"LOSS",-1.8,6.2,-2.5],
    [18,"11:48:24","BEAR",6630.00,6633.25,6628.25,6628.25,"RED",147,1.75,396,-25,5563,52.2,0.58,"C",40,"LOSS",-6.5,1.5,-8.2],
    [19,"11:57:11","BEAR",6634.75,6639.25,6634.25,6634.25,"RED",207,0.5,673,-31,5141,85.5,0.19,"C",35,"LOSS",-7.2,0.8,-10.2],
    [20,"12:07:50","BEAR",6649.75,6652.00,6647.00,6647.00,"RED",325,2.75,435,0,2265,72.1,0.22,"C",40,"WIN",1.8,9.8,-0.5],
    [21,"12:13:24","BEAR",6640.50,6641.75,6636.75,6636.75,"RED",270,3.75,391,-75,5974,47.4,0.22,"C",40,"LOSS",-3.0,5.0,-3.5],
    [22,"12:35:22","BULL",6636.00,6638.25,6633.25,6638.25,"GREEN",-243,2.25,19,-579,7837,39.5,0.06,"C",45,"WIN",1.8,9.8,-0.2],
    [23,"12:42:12","BEAR",6646.50,6648.00,6643.00,6643.00,"RED",381,3.5,576,-1,2957,60.0,0.26,"B",50,"WIN",18.8,26.8,-1.2],
    [24,"12:52:17","BULL",6628.50,6631.50,6626.50,6631.50,"GREEN",-125,3.0,398,-650,8743,55.2,0.68,"B",60,"LOSS",-5.2,2.8,-6.0],
    [25,"13:56:08","BEAR",6634.50,6636.25,6631.25,6631.25,"RED",161,3.25,304,-49,5755,20.6,0.63,"C",45,"LOSS",-3.8,4.2,-5.8],
    [26,"14:13:29","BEAR",6643.75,6647.50,6642.50,6642.50,"RED",107,1.25,385,-20,4851,67.0,0.73,"B",55,"WIN",10.5,18.5,-4.0],
    [27,"14:14:42","BULL",6642.50,6646.50,6641.50,6641.50,"RED",-260,1.0,247,-260,5060,46.4,0.09,"LOG",30,"LOSS",-5.5,2.5,-5.5],
    [28,"14:36:55","BULL",6629.75,6631.50,6626.50,6631.50,"GREEN",-180,1.75,39,-547,4238,41.4,0.66,"C",45,"WIN",17.8,18.2,-7.5],
]

for row_i, row_data in enumerate(signals_data, 2):
    for col_i, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_i, column=col_i, value=val)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Check if user signal
    time_str = row_data[1][:5]
    is_user = time_str in user_times

    # Your Signal? column
    ws.cell(row=row_i, column=22, value="YES" if is_user else "").border = border
    if is_user:
        ws.cell(row=row_i, column=22).fill = user_fill

    # Comment column (yellow, editable)
    cell_comment = ws.cell(row=row_i, column=23, value="")
    cell_comment.border = border
    cell_comment.fill = yellow_fill

    # Color outcome
    outcome = row_data[17]
    if outcome == "WIN":
        ws.cell(row=row_i, column=18).fill = green_fill
        ws.cell(row=row_i, column=19).fill = green_fill
    else:
        ws.cell(row=row_i, column=18).fill = red_fill
        ws.cell(row=row_i, column=19).fill = red_fill

# Auto-width columns
for col in range(1, len(headers) + 1):
    max_len = len(str(headers[col - 1]))
    for row in range(2, len(signals_data) + 2):
        val = ws.cell(row=row, column=col).value
        if val:
            max_len = max(max_len, len(str(val)))
    ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 20)

ws.column_dimensions[get_column_letter(23)].width = 45
ws.freeze_panes = "A2"

wb.save("exports/v4_signals_mar19.xlsx")
print("Saved: exports/v4_signals_mar19.xlsx")
print(f"{len(signals_data)} signals ready for your comments")
