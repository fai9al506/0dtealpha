"""Generate ES Absorption Excel from pre-pulled JSON data."""
import json
from datetime import datetime, timedelta, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

data = json.load(open('abs_data.json'))
print(f"Loaded {len(data)} trades")

wb = Workbook()
ws = wb.active
ws.title = "ES Absorption Signals"

# Styles
hdr_font = Font(bold=True, color="FFFFFF", size=9)
hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
sw1_hdr = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
sw2_hdr = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
trig_hdr = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
sw1_bg = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
sw2_bg = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
trig_bg = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")
win_bg = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
loss_bg = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
exp_bg = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
center = Alignment(horizontal='center', vertical='center')

# Column definitions with group markers
cols = [
    # Trade info (1-10)
    ("ID", "trade"), ("Date", "trade"), ("Time ET", "trade"), ("Direction", "trade"),
    ("Pattern", "trade"), ("Grade", "trade"), ("Score", "trade"),
    ("SPX Spot", "trade"), ("ES Price", "trade"), ("Paradigm", "trade"),
    # Outcome (11-14)
    ("Result", "outcome"), ("PnL", "outcome"), ("MaxProfit", "outcome"), ("MaxLoss", "outcome"),
    # Swing 1 — Reference/Older (15-22)
    ("Sw1 Type", "sw1"), ("Sw1 Time ET", "sw1"), ("Sw1 Bar#", "sw1"),
    ("Sw1 Price", "sw1"), ("Sw1 CVD", "sw1"), ("Sw1 Delta", "sw1"),
    ("Sw1 Volume", "sw1"), ("Sw1 BuyVol", "sw1"),
    # Swing 2 — Recent (23-30)
    ("Sw2 Type", "sw2"), ("Sw2 Time ET", "sw2"), ("Sw2 Bar#", "sw2"),
    ("Sw2 Price", "sw2"), ("Sw2 CVD", "sw2"), ("Sw2 Delta", "sw2"),
    ("Sw2 Volume", "sw2"), ("Sw2 BuyVol", "sw2"),
    # Divergence metrics (31-36)
    ("Price Δ", "div"), ("CVD Δ", "div"), ("CVD Z", "div"),
    ("Price/ATR", "div"), ("VolRatio", "div"), ("ATR", "div"),
    # Trigger bar (37-47)
    ("Trig Bar#", "trig"), ("Trig Time ET", "trig"),
    ("Trig Open", "trig"), ("Trig High", "trig"), ("Trig Low", "trig"), ("Trig Close", "trig"),
    ("Trig Vol", "trig"), ("Trig BuyVol", "trig"), ("Trig SellVol", "trig"),
    ("Trig Delta", "trig"), ("Trig CVD", "trig"),
    # Extra (48-49)
    ("Tier", "extra"), ("Notes", "extra"),
]

# Write headers
for ci, (name, group) in enumerate(cols, 1):
    cell = ws.cell(row=1, column=ci, value=name)
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin
    if group == 'sw1': cell.fill = sw1_hdr
    elif group == 'sw2': cell.fill = sw2_hdr
    elif group == 'trig': cell.fill = trig_hdr
    else: cell.fill = hdr_fill

def parse_to_et(ts_str):
    """Convert UTC timestamp string to ET time string."""
    if not ts_str: return ''
    try:
        # Parse the timestamp
        ts_str = str(ts_str)
        if '+00:00' in ts_str or 'Z' in ts_str:
            # UTC — subtract 5 hours for EST
            dt = datetime.fromisoformat(ts_str.replace('Z', '+00:00'))
            dt_et = dt - timedelta(hours=5)
            return dt_et.strftime('%H:%M:%S')
        elif '-05:00' in ts_str or '-04:00' in ts_str:
            # Already ET
            dt = datetime.fromisoformat(ts_str)
            return dt.strftime('%H:%M:%S')
        else:
            # Assume UTC
            dt = datetime.fromisoformat(ts_str)
            dt_et = dt - timedelta(hours=5)
            return dt_et.strftime('%H:%M:%S')
    except:
        return str(ts_str)[:8]

row = 2
for t in data:
    abs_d = t.get('abs_details', {})
    best = abs_d.get('best_swing', {})
    sw1_info = best.get('ref_swing', {})   # older swing
    sw2_info = best.get('swing', {})       # recent swing
    sw1_bar = t.get('sw1_bar') or {}
    sw2_bar = t.get('sw2_bar') or {}
    trig_bar = t.get('trigger_bar') or {}
    pattern = abs_d.get('pattern', '')

    # Parse trade time to ET
    ts_str = t['ts']
    trade_time_et = parse_to_et(ts_str)
    trade_date = ''
    try:
        dt = datetime.fromisoformat(ts_str)
        dt_et = dt - timedelta(hours=5)
        trade_date = dt_et.strftime('%Y-%m-%d')
    except: pass

    # Swing prices/CVD from abs_details (authoritative), fall back to bar data
    sw1_price = sw1_info.get('price', '') or sw1_bar.get('close', '')
    sw1_cvd = sw1_info.get('cvd', '') if sw1_info.get('cvd') is not None else sw1_bar.get('cvd', '')
    sw1_vol = sw1_info.get('volume', '') or sw1_bar.get('volume', '')
    sw2_price = sw2_info.get('price', '') or sw2_bar.get('close', '')
    sw2_cvd = sw2_info.get('cvd', '') if sw2_info.get('cvd') is not None else sw2_bar.get('cvd', '')
    sw2_vol = sw2_info.get('volume', '') or sw2_bar.get('volume', '')

    # Delta and buy vol from actual bar data
    sw1_delta = sw1_bar.get('delta', '')
    sw1_buyvol = sw1_bar.get('buy_vol', '')
    sw2_delta = sw2_bar.get('delta', '')
    sw2_buyvol = sw2_bar.get('buy_vol', '')

    # Calculate divergence
    price_delta = ''
    cvd_delta = ''
    if sw1_price and sw2_price:
        try: price_delta = round(float(sw2_price) - float(sw1_price), 2)
        except: pass
    if sw1_cvd != '' and sw2_cvd != '':
        try: cvd_delta = round(float(sw2_cvd) - float(sw1_cvd), 1)
        except: pass

    cvd_z = best.get('cvd_z', '')
    price_atr = best.get('price_atr', '')
    vol_ratio = abs_d.get('vol_ratio', '')
    atr = abs_d.get('atr', '')
    tier = abs_d.get('pattern_tier', '')

    # Parse swing times — fall back to bar data ts_start when swing info has no ts
    sw1_time = parse_to_et(sw1_info.get('ts', '') or sw1_bar.get('ts_start', ''))
    sw2_time = parse_to_et(sw2_info.get('ts', '') or sw2_bar.get('ts_start', ''))
    trig_time = parse_to_et(trig_bar.get('ts_start', ''))

    values = [
        t['id'], trade_date, trade_time_et, t['direction'],
        pattern, t['grade'], t['score'],
        t['spot'], t['es_price'], t['paradigm'] or '',
        # Outcome
        t['result'] or 'OPEN', t['pnl'], t['max_profit'], t['max_loss'],
        # Sw1
        sw1_info.get('type', ''), sw1_time, sw1_info.get('bar_idx', ''),
        sw1_price, sw1_cvd, sw1_delta, sw1_vol, sw1_buyvol,
        # Sw2
        sw2_info.get('type', ''), sw2_time, sw2_info.get('bar_idx', ''),
        sw2_price, sw2_cvd, sw2_delta, sw2_vol, sw2_buyvol,
        # Divergence
        price_delta, cvd_delta,
        round(cvd_z, 2) if isinstance(cvd_z, (int, float)) else cvd_z,
        round(price_atr, 2) if isinstance(price_atr, (int, float)) else price_atr,
        round(vol_ratio, 2) if isinstance(vol_ratio, (int, float)) else vol_ratio,
        round(atr, 2) if isinstance(atr, (int, float)) else atr,
        # Trigger bar
        trig_bar.get('bar_idx', abs_d.get('bar_idx', '')),
        trig_time,
        trig_bar.get('open', ''), trig_bar.get('high', ''),
        trig_bar.get('low', ''), trig_bar.get('close', ''),
        trig_bar.get('volume', ''), trig_bar.get('buy_vol', ''),
        trig_bar.get('sell_vol', ''), trig_bar.get('delta', ''),
        trig_bar.get('cvd', ''),
        # Extra
        tier, ''  # Notes column for user to fill
    ]

    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=ci, value=val)
        cell.border = thin
        cell.alignment = center
        grp = cols[ci-1][1]
        if grp == 'sw1': cell.fill = sw1_bg
        elif grp == 'sw2': cell.fill = sw2_bg
        elif grp == 'trig': cell.fill = trig_bg

    # Color result
    res_cell = ws.cell(row=row, column=11)
    if t['result'] == 'WIN': res_cell.fill = win_bg
    elif t['result'] == 'LOSS': res_cell.fill = loss_bg
    elif t['result'] == 'EXPIRED': res_cell.fill = exp_bg

    row += 1

# Column widths
widths = {
    'A': 5, 'B': 11, 'C': 9, 'D': 8, 'E': 18, 'F': 6, 'G': 6,
    'H': 9, 'I': 9, 'J': 14,
    'K': 8, 'L': 6, 'M': 8, 'N': 8,
}
for col_letter, w in widths.items():
    ws.column_dimensions[col_letter].width = w
# Auto-size remaining
for ci in range(15, len(cols)+1):
    letter = ws.cell(row=1, column=ci).column_letter
    ws.column_dimensions[letter].width = 10

# Freeze panes
ws.freeze_panes = 'E2'

# ========== Summary Sheet ==========
ws2 = wb.create_sheet("Guide")
guide_data = [
    ("ES Absorption Signals — ATAS Cross-Check Guide", "", ""),
    ("", "", ""),
    ("COLUMN GROUPS:", "", ""),
    ("White (A-J)", "Trade info — ID, time, direction, pattern, paradigm", ""),
    ("Green (Sw1)", "REFERENCE swing (older) — first in the pair", "Compare with ATAS swing point"),
    ("Blue (Sw2)", "RECENT swing — second in the pair", "Compare with ATAS swing point"),
    ("Orange (Trig)", "TRIGGER bar — the bar that fired the signal", "This is the entry bar"),
    ("", "", ""),
    ("PATTERN MEANINGS:", "", ""),
    ("sell_exhaustion", "Lower low + Higher CVD → BUY", "Sellers pushed price down but CVD rising = exhausted"),
    ("sell_absorption", "Higher low + Lower CVD → BUY", "Price held up while CVD dropped = passive buyers"),
    ("buy_exhaustion", "Higher high + Lower CVD → SELL", "Buyers pushed price up but CVD dropping = exhausted"),
    ("buy_absorption", "Lower high + Higher CVD → SELL", "Price failed while CVD rose = passive sellers"),
    ("", "", ""),
    ("KEY METRICS:", "", ""),
    ("CVD Z", "CVD gap / rolling std dev (20 bars). Higher = stronger divergence", "Min 0.5 to fire"),
    ("Price/ATR", "Price distance between swings / avg bar range (20 bars)", ""),
    ("VolRatio", "Trigger bar volume / 10-bar avg. Must be >= 1.4", "Volume confirmation"),
    ("Tier", "2 = Exhaustion (priority), 1 = Absorption", "Tier 2 beats Tier 1 on conflicts"),
    ("", "", ""),
    ("HOW TO CROSS-CHECK:", "", ""),
    ("1.", "Open ATAS with ES 5-pt range bars", ""),
    ("2.", "Find each trade by Trig Time ET", ""),
    ("3.", "Compare Sw1/Sw2 price, CVD, volume with ATAS candles", ""),
    ("4.", "Check: are the swing points (H/L) actually there on ATAS?", ""),
    ("5.", "Check: does the CVD divergence match what ATAS shows?", ""),
    ("6.", "Add your notes in the Notes column", ""),
    ("", "", ""),
    ("WHAT TO LOOK FOR:", "", ""),
    ("- Missing swings", "Our pivot detection (left=2, right=2) may miss some ATAS pivots", ""),
    ("- CVD mismatch", "Our CVD resets at RTH open, ATAS carries overnight", ""),
    ("- Volume mismatch", "We aggregate Rithmic sub-fills; ATAS may count differently", ""),
    ("- False signals", "Pattern fired but visual context shows no real divergence", ""),
]

for ri, (a, b, c) in enumerate(guide_data, 1):
    ws2.cell(row=ri, column=1, value=a).font = Font(bold=True) if ri <= 1 or a.endswith(':') else Font()
    ws2.cell(row=ri, column=2, value=b)
    ws2.cell(row=ri, column=3, value=c).font = Font(italic=True, color="666666")

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 60
ws2.column_dimensions['C'].width = 45
ws2.cell(row=1, column=1).font = Font(bold=True, size=14)

# Save
out = "ES_Absorption_Signals_v2.xlsx"
wb.save(out)
print(f"Saved {row-2} trades to {out}")
print(f"  - {sum(1 for t in data if t.get('abs_details',{}).get('best_swing'))} with full swing data")
print(f"  - {sum(1 for t in data if not t.get('abs_details',{}).get('best_swing'))} without (pre-rewrite)")
